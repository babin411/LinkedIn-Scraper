import os
import csv
import re
import openpyxl
from openpyxl import Workbook

from time import sleep
from scrapy.spiders import CrawlSpider
from scrapy.selector import Selector
from scrapy.http import Request

from selenium import webdriver

from webdriver_manager.firefox import GeckoDriverManager 

from selenium.webdriver.common.keys import Keys

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC  
from selenium.webdriver.support.ui import WebDriverWait

# from sklearn.feature_extraction.text import CountVectorizer


class FinalScrapeSpider(CrawlSpider):
    name = 'final_scrape'
    allowed_domains = ['www.linkedin.com']
    start_urls = ['http://www.linkedin.com/']


    def __init__(self):
        credentials_path = os.path.join(os.getcwd(), 'credentials.txt')
        file = open(credentials_path)
        lines = file.read().splitlines()

        config_dict = dict()
        for i in range(len(lines)):
            key, value = lines[i].split('=')[0], lines[i].split('=')[1]
            config_dict[key]=value
            
        self.driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
        self.driver.get(self.start_urls[0])
        #clikcing signin button
        self.driver.find_element_by_xpath('//a[@class="nav__button-secondary"]').click()
        username = self.driver.find_element_by_id("username")
        #mimic user typing to send data to the username or email to the form field
        username.send_keys(config_dict['email'])
        sleep(2)
        
        #find password field
        password = self.driver.find_element_by_id('password')
        #mimic user typing password to send password to the form field
        password.send_keys(config_dict['password'])
        sleep(2)
        
        #find submit button
        self.driver.find_element_by_xpath('//div[@class="login__form_action_container "]/button').click()


    def start_requests(self):
        path = os.path.join(os.getcwd(), 'Copy MBA.xlsx')
        wb_obj = openpyxl.load_workbook(path)
        sheet_names = wb_obj.sheetnames
        sheet_obj = wb_obj[sheet_names[2]]
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        # with open('top25_engg.csv', 'a', newline='') as file:
        for i in range(2, max_row+1):
            first_name = sheet_obj.cell(row=i, column=4).value
            last_name = sheet_obj.cell(row=i, column=5).value
            if last_name is not None:
                name = first_name + ' ' + last_name
            else:
                name = first_name
            
            if name is None:
                continue
            
            name = re.sub(r'(\(*Mrs\.\|Mr\.|Dr\.|Prof\.|Er\.|\(.*\)\)*)', '', str(name), flags= re.I)
            name = name.strip()
            institution_object = sheet_obj.cell(row=i, column=1)
            linkedin_object = sheet_obj.cell(row=i,column=9)
            college = institution_object.value
            
            #search field
            if linkedin_object is not None:
                self.driver.find_element_by_xpath('//input[@class="search-global-typeahead__input always-show-placeholder"]').send_keys(name.strip().lower())
                sleep(2)
                self.driver.find_element_by_xpath('//input[@class="search-global-typeahead__input always-show-placeholder"]').send_keys(Keys.ENTER)
                #filters        
                sleep(5)
                # self.driver.find_element_by_xpath('.//button[contains(@aria-label, "People")]').click()
                # WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable(By.XPATH, ".//button[contains(@aria-label, "People")]")).click()
                # WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, './/button[contains(@aria-label, "People")]'))).click()
                wait = WebDriverWait(self.driver, 20)
                button = wait.until(EC.element_to_be_clickable((By.XPATH, './/button[contains(@aria-label, "People")]')))
                # element.click()
                self.driver.execute_script("arguments[0].click()", button)
                
                # sleep(2)
                
                #parse_page
                sel = Selector(text=self.driver.page_source)
                profiles = sel.xpath('//ul[@class="reusable-search__entity-result-list list-style-none"]/li')
                
                if profiles:            
                    for profile in profiles:
                        text = profile.xpath('.//div[contains(@class, "entity-result__primary-subtitle")]/text()').extract()[1]
                        text = text.strip().lower()
                        if college.strip().lower() in text:
                            profile_link = profile.xpath('.//span[@class="entity-result__title-text t-16"]/a[@class="app-aware-link"]/@href').extract_first()
                            self.driver.get(profile_link)
                            self.driver.find_element_by_xpath('//a[contains(@id, "contact-info")]').click()
                            sleep(3)
                            sel = Selector(text=self.driver.page_source)                        
                            linkedin_profile = sel.xpath('//section[contains(@class, "ci-vanity-url")]/div/a/@href').extract_first()
                            try:
                                email = sel.xpath('//section[contains(@class, "ci-email")]/div/a/@href').extract_first()
                            except Exception as e:
                                print(f'Exception as {e}')
                            else:
                                if email is not None:
                                    email = email[7:]
                            try:
                                phone = sel.xpath('//section[contains(@class, "ci-phone")]/ul/li/span/text()').extract_first()
                            except Exception as e:
                                print(f'Exception as {e}')
                            else:
                                if phone is not None:
                                    phone = phone.strip()   
                            # print('----------------------------')
                            # first_name = sheet_obj.cell(row=i, column=4).value
                            sheet_obj.cell(row=i,column=8).value = linkedin_profile
                            sheet_obj.cell(row=i,column=7).value = email
                            sheet_obj.cell(row=i,column=9).value = phone
                            wb_obj.save(path)
                            # print(f'Profile: {linkedin_profile}\nEmail: {email}\nPhone:{phone}')
                            self.driver.find_element_by_xpath('//button[contains(@class, "artdeco-modal__dismiss")]').click()
                            sleep(3)
                            # print('----------------------------')
                        else:
                            self.driver.find_element_by_xpath('//input[@class="search-global-typeahead__input always-show-placeholder"]').clear()
                            sleep(4)
                else:
                    self.driver.find_element_by_xpath('//input[@class="search-global-typeahead__input always-show-placeholder"]').clear()
                    sleep(3)
        

    def parse(self, response):
        pass
